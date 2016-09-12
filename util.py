#!/usr/local/bin/python2.7
# encoding: utf-8
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import time
import math
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors


def loadData(fileName):
    print "start load data %s" % time.clock()
    wholeWorkBook = {}
    QuanData = load_workbook(fileName)
    print "end load data %s" % time.clock()
    sheetNames = QuanData.get_sheet_names()
    #print sheetNames
    sheetNames.sort()   
    for currentSheetName in sheetNames:
        print "load sheet Name %s: %s" % (currentSheetName, time.clock())
        sheetData = QuanData.get_sheet_by_name(currentSheetName)
        rows = sheetData.rows;
        columns = sheetData.columns;
        rowsNum = len(rows)
        columsNum = len(columns)

        cnt = 0
        sheetRowTitleValue = []
        sheetColValue = []       
        #print "Data in ", currentSheetName, ":\n"
        for row in rows:
            for colValue in row:   
                             
                if(cnt < columsNum):
                    sheetRowTitleValue.append(colValue.value)
                else:
                    sheetColValue.append(colValue.value)
                    
                cnt = cnt + 1
        
        wholeColumData = []
        for i in range(columsNum):
            columnData = []
            for j in range(rowsNum - 1):
                index = j * columsNum + i;
                #print index
                columnData.append(sheetColValue[index])
            wholeColumData.append(columnData)
               
        
        #print wholeColumData
        
        sheetDataFinal = {}
        for i in range(columsNum):
            sheetDataFinal[sheetRowTitleValue[i]] = wholeColumData[i]
            
        #print sheetDataFinal
        
        wholeWorkBook[currentSheetName] = sheetDataFinal
        
    return wholeWorkBook




def extractPartData(wordBook, rows = 6, output="partial.xlsx"):
    
    wb = Workbook()
    #remove default worksheet
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    #wb.remove_sheet("Sheet")
    sheetNames = list(wordBook)
    sheetNames.sort()
    
    
    for currentSheetName in sheetNames:
        ws = wb.create_sheet(title=currentSheetName)
        sheetData = wordBook[currentSheetName]
        sheetTitle = list(sheetData)
        
        ws.append(sheetTitle)
        
        print "Data in sheet: ", currentSheetName, ":\n"
        allColumnData = []
        for currentTitle in sheetTitle:
            columnData = sheetData[currentTitle]
            allColumnData.append(columnData)
        
        rowsNum = len(allColumnData[0])
        columnsNum = len(sheetTitle)
        
        print rowsNum, columnsNum
        
        #for debug file quan.xlsx
        rowsNum = rows
        for i in range(columnsNum):
            for j in range(rowsNum):
                ws.cell(row = j + 2, column = i + 1, value=allColumnData[i][j])
        
            #print columnData
    
    
    wb.save(filename=output)
    
    
    
    
    
    
def loadScreenData(fileName):
    print "start load data %s" % time.clock()
    wholeWorkBook = {}
    QuanData = load_workbook(fileName)
    print "end load data %s" % time.clock()
    sheetNames = QuanData.get_sheet_names()
    #print sheetNames
    sheetNames.sort()   
    for currentSheetName in sheetNames:
        print "load sheet Name %s: %s" % (currentSheetName, time.clock())
        sheetData = QuanData.get_sheet_by_name(currentSheetName)
        rows = sheetData.rows;
        columns = sheetData.columns;
        rowsNum = len(rows)
        columsNum = len(columns)

        cnt = 0
        sheetRowTitleValue = []
        sheetColValue = []       
        #print "Data in ", currentSheetName, ":\n"
        for row in rows:
            for colValue in row:   
                             
                if(cnt < columsNum):
                    sheetRowTitleValue.append(colValue.value)
                else:
                    sheetColValue.append(colValue.value)
                    
                cnt = cnt + 1
        
        wholeColumData = []
        for i in range(columsNum):
            columnData = []
            for j in range(rowsNum - 1):
                index = j * columsNum + i;
                #print index
                columnData.append(sheetColValue[index])
            wholeColumData.append(columnData)
               
        
        #print wholeColumData
        
        sheetDataFinal = {}
        for i in range(columsNum):
            sheetDataFinal[sheetRowTitleValue[i]] = wholeColumData[i]
            
        #print sheetDataFinal
        
        wholeWorkBook[currentSheetName] = sheetDataFinal
        
    return wholeWorkBook

def initNewWordBook():
    wb = Workbook()
    #remove default worksheet
    #wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    return wb

def writeWordBook(wordbook, output="newbook.xlsx"):
    wordbook.save(filename = output)
    
def writeDataToColumn(wordBook, sheetName, data, columnIndex):
    ws = wordBook.get_sheet_by_name(sheetName)
    
    for i in range(len(data)):
        ws.cell(row = i + 2, column = columnIndex, value = data[i])
        

def extractMZExpectData(screenDataBook):
       
    mzExpectInfo = {}
    sheetNames = list(screenDataBook)
    compoundNameTitle = "Compound Name"
    mzExpectTitle = "m/z (Expected)"
    for currSheetName in sheetNames:
        sheetData = screenDataBook[currSheetName]
        compoundNameData = sheetData[compoundNameTitle]
        mzExpectData = sheetData[mzExpectTitle]
        
        for i in range(len(compoundNameData)):
            mzExpectInfo[compoundNameData[i]] = mzExpectData[i]
            
    
    return mzExpectInfo
        
        
    