# encoding: utf-8
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from Logger import MyLogger
import logging
import math
import re
import pprint
from subprocess import Popen
from popen2 import popen4
import glob
import time
#import operator

logger = MyLogger("Lipid-Logger", logging.INFO).getLogger()
ms2Window = 0.2
topRTRange = 0.25
#保存合并后的名字与原始名字的映射
#其中，key是合并后的新名字，val是合并前的名字，可能有多个，因为多个不同的名字经过合并后是相同的
lipidAlias = {}
def nonBlankLines(f):
    for l in f:
        line = l.rstrip()
        if line and line[0] != '#':
            yield line

def readTextFile(fileName): 
    content = []  
    with open(fileName) as f:
        for line in nonBlankLines(f):
            content.append(line)
    
    return content

def loadTextFile(fileName):
    content = readTextFile(fileName)
    headerLine = []
    if(len(content) > 0):
        headerLine = content[0].split('\t')
    
    headers = []
    dataDict = {}
    for header in headerLine:
        headers.append(header)
        dataDict[header] = {}
    
    
    count = 0   
    #content中的每个元素是文件中的一行，每个元素就是个普通的字符串，这个字符串是文件中的完整一行 
    for item in content:
        item = item.split('\t')
        count += 1
        itemLen = len(item)
        #print count, ":", itemLen, ":" , item
        logger.debug("%d:%d:%s" % (count, itemLen, str(item)))
    #transpose the list of lists
    #content2是一个list，其的每一个元素也是一个list，每个元素的list中包含文件中一行的内容，每个元素list中的每一个元素是该行中的某一列对应的内容，简单来说：content2中的内容就是将文件中的
    #内容按照矩阵元素的方式进行存储了。
    content2 = []
    for item in content:
        item = item.split('\t')
        content2.append(item)
        
    
    #Transpose the content matrix
    logger.debug("file name =" + fileName)
    #第二个for循环先固定每一列的index，然后在内层第一个循环中取content2中的每一行，取每一行的第index个元素作为contentT中第index个元素，从而达到矩阵转置的目的
    
    #调试这个文件：LP08-136-3-pos-1ul-C1-10.txt，他的252行有些问题，只有33列的数据，正常应该是34列的数据，这个文件中第252行的QuantInfo这列木有数据
    logger.debug("Total Columns = " + str(len(content2[0]) - 1))
    cnt = 0
    lenCol = len(content2[0]) - 1
    for x in content2:
        cnt = cnt + 1
        lenx = len(x)
        if(lenx < lenCol):
            logger.debug("<=======>")
            logger.debug("lenx = " + str(lenx) + ", lenCol = " + str(lenCol))
            logger.debug("cnt = " + str(cnt))
            logger.debug("line = " + str(x))
            logger.debug("<=======>")
        
    
    #有些文件，如LP08-136-3-pos-1ul-C1-10.txt的最后一列quantinfo没有数据，这列的数据由于不要提取出来，为处理方便，就不要这列的数据了    
    contentT = [[x[i] for x in content2] for i in range(len(content2[0]) - 1)]
    count = 0
    logger.debug("After transpose:")
    for item in contentT:        
        count += 1
        itemLen = len(item)
        #print count, ":", itemLen, ":" , item
        logger.debug("%d:%d:%s" % (count, itemLen, str(item)))
    
    for item in contentT:
        header = item[0]
        item.pop(0)
        dataDict[header] = item
        
    
    #print "Value in dataDict: ", dataDict
    #printDict(dataDict)
    #print dataDict
        
    return dataDict

def getFileList(path):
    
    finalPath = path + "./*.txt"
    fileList = glob.glob(finalPath)
    
    return fileList
    
#这个函数返回一个dict，key是文件明，value是文件的内容，文件的内容又是一个dict，该dict的key是列的标题，value是该列对应的内容    
def loadAllTextFile(dirName, logger):
    fileList = getFileList(dirName)
    allSampleDataBook = {}
    for f in fileList:
        currSampleData = loadTextFile(f)
        allSampleDataBook[f] = currSampleData 
    
    return allSampleDataBook
def makeTestData():
    dataBook = {}
    fileBook1 = {}
    fileBook2 = {}
    
    fileBook3 = {}
    fileBook4 = {}
    fileBook5 = {}
    fileBook6 = {}
    fileBook7 = {}
    
    
    fileBook1["LipidIon"] = ["ChE(18:3)+H", "ChE(20:5)+NH4", "ChE(20:5)+H", "ChE(22:4)+NH4", "ChE(20:5)+H"]
    fileBook1["Rt"] = ["10.1", "10.2", "10.3","10.4", "11.80"]
    fileBook1["TopRT"] = ["10.25", "10.42", "10.45", "10.88", "11.64"]
    fileBook1["Formula"] = ["fm1", "fm2", "fm2", "fm2", "fm2"]
    fileBook1["Grade"] = ["A", "A", "B", "B", "D"]
    fileBook1["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5"]
    fileBook1["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5"]
    fileBook1["m-Score"] = ["10.25", "10.42", "10.45", "10.88", "11.64"]
    
    fileBook2["LipidIon"] = ["NE(18:3)+H", "ChE(20:5)+NH4", "NE(20:5)+H", "ChE(22:4)+NH4", "ChE(20:5)+H"]
    fileBook2["Rt"] = ["10.1", "10.2", "10.3","10.4", "10.4"]
    fileBook2["TopRT"] = ["10.25", "10.42", "10.45", "10.92", "10.55"]
    fileBook2["Formula"] = ["fm1", "fm2", "fm2", "fm2", "fm2"]
    fileBook2["Grade"] = ["A", "D", "C", "D", "B"]
    fileBook2["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5"]
    fileBook2["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5"]
    fileBook2["m-Score"] = ["10.25", "10.42", "10.45", "10.92", "10.55"]
    
    fileBook3["LipidIon"] = ["TG(4:0/16:0/22:5)+NH4", "TG(4:0/22:5/16:0)+NH4", "DG(16:0/4:0/22:5)+NH4", "DG(4:0/16:0/22:5)+NH4", "TG(4:0/16:0/28:4)+NH4"]
    fileBook3["Rt"] = ["10.1", "10.2", "10.3","10.4", "10.5"]
    fileBook3["TopRT"] = ["10.1", "10.2", "10.3", "10.4", "10.5"]
    fileBook3["Formula"] = ["fm2", "fm2", "fm2", "fm2", "fm2"]
    fileBook3["Grade"] = ["A", "D", "C", "D", "B"]
    fileBook3["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5"]
    fileBook3["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5"]
    fileBook3["m-Score"] = ["10.1", "10.2", "10.3", "10.4", "10.5"]
    
    fileBook4["LipidIon"] = ["PC(16:0e/22:5)+NH4", "PC(22:5/16:0e)+NH4", "SM(16:0/4:0p)+NH4", "SM(4:0p/16:0)+NH4", "PC(d16:0/28:4+O)+NH4", "SM(16:0/4:0p)+NH4"]
    fileBook4["Rt"] = ["10.1", "10.2", "10.3","10.4", "10.5", "10.8"]
    fileBook4["TopRT"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8"]
    fileBook4["Formula"] = ["fm3", "fm3", "fm3", "fm3", "fm3", "fm3"]
    fileBook4["Grade"] = ["A", "D", "C", "D", "B", "A"]
    fileBook4["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5", "OM6"]
    fileBook4["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5", "Area6"]
    fileBook4["m-Score"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8"]
    
    fileBook5["LipidIon"] = ["Cer(16:0e/22:5)+NH4", "Cer(22:5e/16:0)+NH4", "phSM(16:0/4:0p)+NH4", "phSM(4:0p/16:0)+NH4", "Cer(16:0e/22:5)+NH4", "phSM(16:0/4:0p)+NH4", "phSM(16:0/4:0p)+NH4"]
    fileBook5["Rt"] = ["10.1", "10.2", "10.3","10.4", "10.5", "10.8", "10.4"]
    fileBook5["TopRT"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8", "10.4"]
    fileBook5["Formula"] = ["fm5", "fm5", "fm5", "fm5", "fm5", "fm5", "fm5"]
    fileBook5["Grade"] = ["A", "B", "B", "A", "B", "A", "B"]
    fileBook5["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5", "OM6", "OM7"]
    fileBook5["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5", "Area6", "Area7"]
    fileBook5["m-Score"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8", "10.4"]
    
    #注意：不可能出现两盒化合物名称相同，但是formula不同的情况，在构造测试数据的时候特别注意
    fileBook6["LipidIon"] = ["CerG1(16:0e/22:5)+NH4", "CerG1(22:5e/16:0)+NH4", "phSMG2(16:0/4:0p)+NH4", "phSMG2(4:0p/16:0)+NH4", "CerG1(16:0e/22:5)+NH4", "phSMG2(16:0/4:0p)+NH4", "phSMG2(16:0/4:0p)+NH4"]
    fileBook6["Rt"] = ["10.1", "10.2", "10.3","10.4", "10.5", "10.8", "10.4"]
    fileBook6["TopRT"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8", "10.4"]
    fileBook6["Formula"] = ["fm6", "fm6", "fm6", "fm6", "fm6", "fm6", "fm6"]
    fileBook6["Grade"] = ["C", "D", "D", "C", "D", "C", "D"]
    fileBook6["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5", "OM6", "OM7"]
    fileBook6["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5", "Area6", "Area7"]
    fileBook6["m-Score"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8", "10.4"]
    
    
    fileBook7["LipidIon"] = ["CerG2(16:0e/22:5)+NH4", "CerG2(22:5e/16:0)+NH4", "phSMG3(16:0/4:0p)+NH4", "phSMG3(4:0p/16:0)+NH4", "CerG2(16:0e/22:5)+NH4", "phSMG3(16:0/4:0p)+NH4", "phSMG3(4:0p/16:0)+NH4"]
    fileBook7["Rt"] = ["10.1", "10.2", "10.3","10.4", "10.5", "10.8", "10.4"]
    fileBook7["TopRT"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8", "10.4"]
    fileBook7["Formula"] = ["fm7", "fm7", "fm7", "fm7", "fm7", "fm7", "fm7"]
    fileBook7["Grade"] = ["A", "D", "D", "C", "C", "B", "A"]
    fileBook7["ObsMz"] = ["OM1", "OM2", "OM3", "OM4", "OM5", "OM6", "OM7"]
    fileBook7["Area"] = ["Area1", "Area2", "Area3", "Area4", "Area5", "Area6", "Area7"]
    fileBook7["m-Score"] = ["10.1", "10.2", "10.3", "10.4", "10.5", "10.8", "10.4"]
    
    
    dataBook["f1"] = fileBook1;
    dataBook["f2"] = fileBook2;
    dataBook["f3"] = fileBook3;
    dataBook["f4"] = fileBook4;
    dataBook["f5"] = fileBook5;
    dataBook["f6"] = fileBook6;
    dataBook["f7"] = fileBook7;
    
    
    return dataBook        

#将样品数据组织成五元组的形式(lipidIons, Rts, TopRTs, diff, Formulas, Grades)        
def makeTuple(dataBook):
    logger.debug("Start to make databook as tuple list")    
    lipidInfo = []    
    files = list(dataBook)
    for f in files:
        fileData = dataBook[f]
        lipidIons = fileData["LipidIon"]
        Rts = fileData["Rt"]
        TopRTs = fileData["TopRT"]
        Formulas = fileData["Formula"]
        Grades = fileData["Grade"]
        ObsMz = fileData["ObsMz"]
        Area = fileData["Area"]
        mScore = fileData["m-Score"]
        
        logger.debug("lipidIons:" + str(lipidIons)) 
        logger.debug("Rts:" + str(Rts))
        logger.debug("TopRTs:" + str(TopRTs))
        logger.debug("Formulas:" + str(Formulas))
        logger.debug("Grades:" + str(Grades))
        logger.debug("obsMz:" + str(ObsMz))
        logger.debug("--------------------------------------------------------------")
        
        columnDataSize = len(lipidIons)
        for i in range(columnDataSize):
            diff = math.fabs(float(Rts[i]) - float(TopRTs[i]))
            li = (lipidIons[i], Rts[i], TopRTs[i], diff, Formulas[i], Grades[i], ObsMz[i], Area[i], mScore[i], f)
            lipidInfo.append(li)
    
    #lipidInfo = sorted(lipidInfo,key=operator.itemgetter(3,0),reverse=False)            
    logger.debug("End to make databook as tuple list")    
    return lipidInfo     

#在lipidInfo中删除RT和TopRT差值大于0.2的元组
def rm0dot2(lipidInfo):
    
    logger.debug("Start to remove item that the gap between RT and TopRT is bigger than ms2Window(0.2)")
    newLipidInfo = []
    for item in lipidInfo:
        diff = item[3]
        if(diff <= ms2Window):
            newLipidInfo.append(item)
    
    logger.debug("End to remove item that the gap between RT and TopRT is bigger than ms2Window(0.2)")
    
    return newLipidInfo


#计算每一种化合物在所有文件中的TopRT的平均值
def calTopRTAvg(lipidInfo):
    avgTable = {};
    topRTTable = {}
    for item in lipidInfo:
        lipidName = item[0]
        topRT = item[2]
        
        if(topRTTable.has_key(lipidName)):
            topRTTable[lipidName].append(float(topRT))
        else:
            topRTTable[lipidName] = [float(topRT)]


    logger.debug("topRT Table = " + str(topRTTable))
    
    for(k, v) in topRTTable.iteritems():
        if(len(v) == 0):
            avgTable[k] = 0.0
        else:
            avgTable[k] = sum(v)/len(v)
        
    logger.debug("avg Table = " + str(avgTable))
    return (avgTable, topRTTable)

#拿到所有相同formula对应的set(化合物),具有相同formula的化合物可以通过topRTTable这个dict来取得其所有的topRT值
def getFormulaMap(lipidInfo):
    
    logger.debug("Start to get lipid names with same formula")
    #保存f->c的映射
    f2c = {}
    #保存c->f的映射
    c2f = {}
    
    
    
    for item in lipidInfo:
        formula = item[4]
        lipidName = item[0]
        #topRT = item[2]
        if(f2c.has_key(formula)):            
            f2c[formula].add(lipidName)            
        else:
            f2c[formula] = set([lipidName])
        
        if(c2f.has_key(lipidName)):
            c2f[lipidName].add(formula)
        else:
            c2f[lipidName] = set([formula])
            
    
    logger.debug("end to get lipid names with same formula")
    return (f2c,c2f)

def getObsMZMap(lipidInfo):
    
    logger.debug("Start to get lipid name mapping to obsmz info")
    #保存c->f的映射
    c2om = {}    
    for item in lipidInfo:
        obsMZ = item[6]
        lipidName = item[0]
       
        
        if(c2om.has_key(lipidName)):
            c2om[lipidName].add(obsMZ)
        else:
            c2om[lipidName] = set([obsMZ])
            
    
    logger.debug("End to get lipid name mapping to obsmz info")
    return c2om

def getC2MScore(lipidInfo):
    
    logger.debug("Start to get lipid name mapping to m-score info")
    #保存c->m-score的映射
    c2ms = {}    
    for item in lipidInfo:
        mScore = item[8]
        lipidName = item[0]
       
        
        if(c2ms.has_key(lipidName)):
            c2ms[lipidName].add(mScore)
        else:
            c2ms[lipidName] = set([mScore])
            
    
    for(k, v) in c2ms.iteritems():
        c2ms[k] = list(v)
    logger.debug("End to get lipid name mapping to m-score info")
    return c2ms

def p5(finalLipidInfo,c2ms):
    
    c2msFinal = {}
    c2MaxMSFinal = {}
    
    for item in finalLipidInfo:
        lipidName = item[0]
        
        mScore = []
        if(c2ms.has_key(lipidName)):
            mScore = c2ms[lipidName]
        else:
            if(lipidAlias.has_key(lipidName)):
                otherNames = lipidAlias[lipidName]
                
                for n in otherNames:
                    mScore = mScore + c2ms[n]
                    
        
        
        
        c2msFinal[lipidName] = mScore
        
        mScore2 = [float(x) for x in mScore]
        c2MaxMSFinal[lipidName] = max(mScore2)
        
    
    return (c2msFinal, c2MaxMSFinal)
    
               
#获取化合物括号部分数字对的个数。
def getGroupNum(compndName):
    
    parenthesesPart = re.search('\(.*\)', compndName).group();   
    #print("parenthesesPart = " + str(parenthesesPart))
    semicolon = re.findall(':', str(parenthesesPart))  
    
    #print("semicolon = " + str(semicolon))    
    
    return len(semicolon)

#重新对某个固定的化合物compndName，其没有合并toprt值进行再次合并：
#将toprt分成两部分，高于newAvg的部分，挑选出来后重新合并，低于newAvg的部分也挑选出来，重新合并
#合并的方式是：重新计算高于和低于newAvg部分的平均值，然后将高于和低于部分的每个元素和算出的新平均值比较，如果
#误差小于topRTRange，则合并，大于的话，单独列出
def reCalUnMergeVals(vldLipidInfo, compndName, newAvg, unMergeVals):    
    smallerValues = []
    biggerValues = []
    
    for v in unMergeVals:
        if(v < newAvg):
            smallerValues.append(v)
        elif(v > newAvg):
            biggerValues.append(v) 

    logger.info("smallerValues = " + str(smallerValues))
    logger.info("biggerValues = " + str(biggerValues))
    smallerAvg = 0.0    
    if(len(smallerValues) >= 1):
        smallerAvg = sum(smallerValues) / len(smallerValues)
    
    logger.info("smallerAvg = " + str(smallerAvg))
    biggerAvg = 0.0
    if(len(biggerValues) >= 1):
        biggerAvg = sum(biggerValues) / len(biggerValues)
    logger.info("biggerAvg = " + str(biggerAvg))    
    smallerMergeVals = []
    smallerUnMergeVals = []
    for v in smallerValues:
        diff = math.fabs(v - smallerAvg)
        if(diff <= topRTRange):
            smallerMergeVals.append(v)
        else:
            smallerUnMergeVals.append(v)
    
    logger.info("smallerMergeVals = " + str(smallerMergeVals))
    logger.info("smallerUnMergeVals = " + str(smallerUnMergeVals))
    
    biggerMergeVals = []
    biggerUnMergeVals = []
    for v in biggerValues:
        diff = math.fabs(v - biggerAvg)
        if(diff <= topRTRange):
            biggerMergeVals.append(v)
        else:
            biggerUnMergeVals.append(v)
            
    logger.info("biggerMergeVals = " + str(biggerMergeVals))
    logger.info("biggerUnMergeVals = " + str(biggerUnMergeVals))
    newSmallerAvg = 0.0
    if(len(smallerMergeVals) > 1):
        newSmallerAvg = sum(smallerMergeVals) / len(smallerMergeVals)
        item = (compndName, newSmallerAvg)
        vldLipidInfo.append(item)
    
    logger.info("newSmallerAvg = " + str(newSmallerAvg))
    newBiggerAvg = 0.0
    if(len(biggerMergeVals) > 1):
        newBiggerAvg = sum(biggerMergeVals) / len(biggerMergeVals)
        item = (compndName, newBiggerAvg)
        vldLipidInfo.append(item)
        
    logger.info("newBiggerAvg = " + str(newBiggerAvg))
    
    for v in smallerUnMergeVals:
        vldLipidInfo.append((compndName, v))
    
    for v in biggerUnMergeVals:
        vldLipidInfo.append((compndName, v))
#提取符合条件2.1的(化合物,toprt)，同时将该化合物信息从候选的lipidInfo中剔除
def p2dot1(f2c, c2TopRTAvg, c2TopRT):
    
    
    logger.debug("Start to process lipid info with only one group num")
    vldLipidInfo = []
    
    
    for (k, v) in f2c.iteritems():
        lipidNames = list(v)
        for n in lipidNames:
            grpNum = getGroupNum(n)
            if(grpNum > 1):
                continue
            
            #处理只有1组数字对的情况
            lipidAvg = c2TopRTAvg[n]
            topRTs = c2TopRT[n]
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((n, newAvg))
            
            #对于与均值的误差大于TopRTRange的值，分成2部分处理，一部分是小于新的均值的部分，一部分是大于新的均值的部分
            #对于小于新的均值的部分，重新计算平均值v，然后计算小于newAvg部分的toprt值 与v的差diff2，如果diff2还是小于topRTRange，则合并，如果大于的话，则不合并
            #对于大于新的均值的部分L2，重新计算L2的平均值V2，然后计算出L2中每个值与V2的差diff3，如果diff3小于topRTRange，则合并，如果大于的话，则不合并，单独列出。
            for value in unMergeValues:
                vldLipidInfo.append((n, value))
                    
                    
    logger.debug("End to process lipid info with only one group num")                
    return vldLipidInfo


    

#提取符合条件2.1的(化合物,toprt)，同时将该化合物信息从候选的lipidInfo中剔除
def p2dot1_ex(f2c, c2TopRTAvg, c2TopRT):
    
    
    logger.debug("Start to process lipid info with only one group num")
    vldLipidInfo = []
    
    
    for (k, v) in f2c.iteritems():
        lipidNames = list(v)
        for n in lipidNames:
            grpNum = getGroupNum(n)
            if(grpNum > 1):
                continue
            
            #处理只有1组数字对的情况
            lipidAvg = c2TopRTAvg[n]
            topRTs = c2TopRT[n]
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            
            newAvg = 0.0;
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((n, newAvg))
            
            #对于与均值的误差大于TopRTRange的值，分成2部分处理，一部分是小于新的均值的部分，一部分是大于新的均值的部分
            #对于小于新的均值的部分，重新计算平均值v，然后计算小于newAvg部分的toprt值 与v的差diff2，如果diff2还是小于topRTRange，则合并，如果大于的话，则不合并
            #对于大于新的均值的部分L2，重新计算L2的平均值V2，然后计算出L2中每个值与V2的差diff3，如果diff3小于topRTRange，则合并，如果大于的话，则不合并，单独列出。
            for value in unMergeValues:
                vldLipidInfo.append((n, value))
                    
                    
    logger.debug("End to process lipid info with only one group num")                
    return vldLipidInfo       
        
#拿到化合物括号中的数字对等信息
def getInfoInPare(lipidName):
    parenthesesPart = re.search('\(.*\)', lipidName).group();
    parenthesesPart = str(parenthesesPart)
    parenthesesPart = parenthesesPart[1:len(parenthesesPart) - 1]    
    logger.debug("parenthesesPart = " + parenthesesPart)    
    elems = re.split('\+|/', parenthesesPart)    
    logger.debug("elems = " + str(elems))
    
    elems = sorted(elems)
    return elems 

#将化合物括号部分的数字对按照升序重新组合
def renameLipid(lipidName):
    prefix = lipidName[0:2]    
    elems = getInfoInPare(lipidName)
    newPare = '/'.join(elems)
    newPare = "(" + newPare + ")"
    suffix = lipidName[lipidName.find(')') + 1:]
    newLipidName = prefix + newPare + suffix
    
    
    #只要被调用一次，就需要保存一次新名字到原始名字的映射
    if(lipidAlias.has_key(newLipidName)):
        lipidAlias[newLipidName].add(lipidName)
    else:
        lipidAlias[newLipidName] = set([lipidName])
    
    return newLipidName
      
#将DG，TG开头的化合物中间的数字部分重新按照升序排列，以便于发现相同的化合物
def reorderDGTG(c2TopRT):
    
    logger.debug("Start to reorder the lipid-->toprts pair with the lipid name start with DG, TG")
    #newc2TopRT中，key保存以DG，TG开头，括号部分经过排序的化合物，value是化合物名称经过处理后是相同的化合物的所有的topRT值
    newc2TopRT = {}
    newc2TopRTAvg = {}
    for (k,v)in c2TopRT.iteritems():
        prefix = k[0:2]
        if(prefix != "DG" and prefix != "TG"):            
            continue;
        
        topRTs = c2TopRT[k]
        
#         elems = getInfoInPare(k)
#         newPare = '/'.join(elems)
#         newPare = "(" + newPare + ")"
#         suffix = k[k.find(')') + 1:]
#         newLipidName = prefix + newPare + suffix

        newLipidName = renameLipid(k)
        
        if(newc2TopRT.has_key(newLipidName)):
            for topRT in topRTs:
                newc2TopRT[newLipidName].append(topRT)
        else:   
                newc2TopRT[newLipidName] = topRTs
        
    
    for (k,v) in newc2TopRT.iteritems():
        if(len(v) == 0):
            newc2TopRTAvg[k] = 0.0
        else:
            newc2TopRTAvg[k] = sum(v)/len(v)
            
    
    logger.debug("End to reorder the lipid-->toprts pair with the lipid name start with DG, TG")
    return (newc2TopRTAvg, newc2TopRT)
        
            
        
                
def p2dot2(f2c, c2TopRT):
    
    logger.debug("Start to process lipid info with lipid name start with DG, TG")
    vldLipidInfo = []
    (newc2TopAvg, newc2TopRT) = reorderDGTG(c2TopRT)
    for(k,v) in f2c.iteritems():
        lipidNames = list(v)
        
        for n in lipidNames:
            prefix = n[0:2]
            #只处理DG或者TG开头的
            if(prefix != "DG" and prefix != "TG"):
                continue
            newLipidName = renameLipid(n)
            
            topRTs = newc2TopRT[newLipidName]
            lipidAvg = newc2TopAvg[newLipidName]
            
#             print("newlipidName" + newLipidName)
#             print("topRTs" + str(topRTs))
#             print("lipidAvg" + str(lipidAvg))
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((newLipidName, newAvg))
            
            for value in unMergeValues:
                vldLipidInfo.append((newLipidName, value))
            
            
            
    logger.debug("End to process lipid info with lipid name start with DG, TG")
    vldLipidInfo = set(vldLipidInfo) #用set去除重复的部分
    vldLipidInfo = list(vldLipidInfo)#再转回list
    return vldLipidInfo        

def reduceName(lipidName):
    alpha = ' dep'
    parenthesesPart = re.search(r'(\(.*\))', lipidName).group()
    parenthesesPart = parenthesesPart[1:-1]  
    
    elems = re.split(':|/|e|d|p|\+', parenthesesPart)
    
    elems = [item for item in elems if item != '']
    
    reduceParePart = str(int(elems[0]) + int(elems[2])) + ":" + str(int(elems[1]) + int(elems[3]))
    
    flag = ''
    for i in range(1, len(alpha)):
        index = parenthesesPart.find(alpha[i])
        if(index != -1):
            flag = alpha[i]    
            break
            
    reduceParePart = reduceParePart + flag
    
    if(len(elems) == 5):
        reduceParePart = reduceParePart + "+" +  elems[-1]
        
    p1 = lipidName.find('(')
    p2 = lipidName.find(')')
    
    newLipidName = lipidName[0:p1+1] + reduceParePart + lipidName[p2:]
    
    
    #只要被调用一次，就需要保存一次新名字到原始名字的映射
    if(lipidAlias.has_key(newLipidName)):
        lipidAlias[newLipidName].add(lipidName)
    else:
        lipidAlias[newLipidName] = set([lipidName])
    
    
    
    return newLipidName


def reorderP_SM(c2TopRT):
    
    logger.debug("Start to reorder the lipid-->toprts pair with the lipid name start with P and SM")
    #newc2TopRT中，key保存以DG，TG开头，括号部分经过排序的化合物，value是化合物名称经过处理后是相同的化合物的所有的topRT值
    newc2TopRT = {}
    newc2TopRTAvg = {}
    for (k,v)in c2TopRT.iteritems():
        prefix = k[0:k.find('(')]
        if(prefix[0] != "P" and prefix != "SM"):            
            continue;
        
        topRTs = c2TopRT[k]
        newLipidName = reduceName(k)
        
        if(newc2TopRT.has_key(newLipidName)):
            for topRT in topRTs:
                newc2TopRT[newLipidName].append(topRT)
        else:   
                newc2TopRT[newLipidName] = topRTs
        
    
    for (k,v) in newc2TopRT.iteritems():
        if(len(v) == 0):
            newc2TopRTAvg[k] = 0.0
        else:
            newc2TopRTAvg[k] = sum(v)/len(v)
            
    
    logger.debug("End to reorder the lipid-->toprts pair with the lipid name start with P and SM")
    return (newc2TopRTAvg, newc2TopRT)    


def p2dot3(f2c, c2TopRT):
    
    logger.debug("Start to process lipid info with lipid name start with DG, TG")
    vldLipidInfo = []
    (newc2TopAvg, newc2TopRT) = reorderP_SM(c2TopRT)
    for(k,v) in f2c.iteritems():
        lipidNames = list(v)
        
        for n in lipidNames:
            prefix = n[0:n.find('(')]
            #只处理DG或者TG开头的
            if(prefix[0] != "P" and prefix != "SM"):
#                 print("continue: prefix = " + prefix)
                continue
            
#             print("prefix = " + prefix)
            newLipidName = reduceName(n)
            
            topRTs = newc2TopRT[newLipidName]
            lipidAvg = newc2TopAvg[newLipidName]
            
#             print("newlipidName: " + newLipidName)
#             print("topRTs: " + str(topRTs))
#             print("lipidAvg: " + str(lipidAvg))
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((newLipidName, newAvg))
            
            for value in unMergeValues:
                vldLipidInfo.append((newLipidName, value))
            
            
            
    logger.debug("End to process lipid info with lipid name start with P, SM")
    vldLipidInfo = set(vldLipidInfo) #用set去除重复的部分
    vldLipidInfo = list(vldLipidInfo)#再转回list
    return vldLipidInfo   


def rmLipidNameIn2dot123(f2c,c2TopRTAvg,c2TopRT, c2Grade):
    
    for(k,v) in f2c.iteritems():
        lipidNames = list(v)
        nv = []
        for n in lipidNames:
            if(n[0] != "P" and getGroupNum(n) != 1 and n[0:n.find('(')] != "SM" and n[0:n.find('(')] != "DG" and n[0:n.find('(')] != "TG"):
                nv.append(n)
        
        f2c[k] = nv
        
    lipidNames = c2TopRT.keys()    
    for n in lipidNames:
        if(n[0] == "P" or getGroupNum(n) == 1 or n[0:n.find('(')] == "SM" or n[0:n.find('(')] == "DG" or n[0:n.find('(')] == "TG"):
            c2TopRT.pop(n)
            
    lipidNames = c2TopRTAvg.keys()    
    for n in lipidNames:
        if(n[0] == "P" or getGroupNum(n) == 1 or n[0:n.find('(')] == "SM" or n[0:n.find('(')] == "DG" or n[0:n.find('(')] == "TG"):
            c2TopRTAvg.pop(n)
                                                 

    lipidNames = c2Grade.keys()    
    for n in lipidNames:
        if(n[0] == "P" or getGroupNum(n) == 1 or n[0:n.find('(')] == "SM" or n[0:n.find('(')] == "DG" or n[0:n.find('(')] == "TG"):
            c2Grade.pop(n)

def getGradeMap(lipidInfo):
    
    c2Grade = {}
    
    for item in lipidInfo:
        lipidName = item[0]
        grade = item[5]
        
        if(c2Grade.has_key(lipidName)):
            c2Grade[lipidName].append(grade)
        else:
            c2Grade[lipidName] = [grade]
            

    
    return c2Grade

def isAllAB(grades):
    flag = True
    
    for g in grades:
        if(g != 'A' and g != 'B'):
            flag = False
            break
    
    return flag

def isAllCD(grades):
    flag = True
    
    for g in grades:
        if(g != 'C' and g != 'D'):
            flag = False
            break
    
    return flag

def isAllABCD(grades):
    flagAB = False
    flagCD = False
    
    for g in grades:
        if(g == 'A' or g == 'B'):
            flagAB = True;
            break
    
    for g in grades:
        if(g == 'C' or g == 'D'):
            flagCD = True;
            break
        
    if(flagAB and flagCD):
        return True
    else:
        return False
    
        

def p2dot4a(f2c, c2TopRTAvg, c2TopRT, c2Grade):
    
    logger.debug("Start to process lipid with grade values as A or B")
    vldLipidInfo = []
    for(k, v) in f2c.iteritems():
        lipidNames = list(v)
        
        for n in lipidNames:
            grades = c2Grade[n]
            flag = isAllAB(grades)
            
            if(flag == False):
                continue
            
            topRTs = c2TopRT[n]
            lipidAvg = c2TopRTAvg[n]
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((n, newAvg))
            
            for value in unMergeValues:
                vldLipidInfo.append((n, value))
                
                
    
    logger.debug("End to process lipid with grade values as A or B")
    
    return vldLipidInfo

def reorderAllGradeCD(c2TopRT, c2Grade):
    newc2TopRTAvg = {}
    newc2TopRT = {}
    
    for (n, v) in c2TopRT.iteritems():
        grades = c2Grade[n]
        flag = isAllCD(grades)
       
        if(flag == False):
            continue 
        newLipidName = reduceName(n)
        
        if(newc2TopRT.has_key(newLipidName)):
            for item in v:
                newc2TopRT[newLipidName].append(item)
        else:
            newc2TopRT[newLipidName] = v
            
    
    for (n, v) in newc2TopRT.iteritems():
        if(len(v) == 0):
            newc2TopRTAvg[n] = 0.0
        else:
            newc2TopRTAvg[n] = sum(v) / len(v)
            
    
    return (newc2TopRTAvg, newc2TopRT)


            
    
    

def p2dot4b(f2c, c2TopRTAvg, c2TopRT, c2Grade):
    logger.debug("Start to process lipid with grade values as C or D")
    vldLipidInfo = []
    
    (newc2TopRTAvg, newc2TopRT) = reorderAllGradeCD(c2TopRT, c2Grade)
    
    #print("newc2TopRT = " + str(newc2TopRT))
    
    for (k, v) in f2c.iteritems():
        lipidNames = list(v)
        
        for n in lipidNames:
            grades = c2Grade[n]
            flag = isAllCD(grades)
            
            if(flag == False):
                continue
            
            newLipidName = reduceName(n)
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            
            topRTs = newc2TopRT[newLipidName]
            lipidAvg = newc2TopRTAvg[newLipidName]
            
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((newLipidName, newAvg))
            
            for value in unMergeValues:
                vldLipidInfo.append((newLipidName, value))
            
            
            
    logger.debug("End to process lipid with grade values as C or D")
    vldLipidInfo = set(vldLipidInfo) #用set去除重复的部分
    vldLipidInfo = list(vldLipidInfo)#再转回list
  
    return vldLipidInfo               


def reorderAllGradeABCD(c2TopRT, c2Grade):
    newc2TopRTAvg = {}
    newc2TopRT = {}
    
    #保存名字经过合并前的原名
    alias = {}
    
    for (n, v) in c2TopRT.iteritems():
        grades = c2Grade[n]
        flag = isAllABCD(grades)
        #print("flag = " +  str(flag) + ", name = " + n + ", grades " + str(grades))
        if(flag == False):
            continue 
        newLipidName = reduceName(n)
        
        #newLipidName = n + "/" + newLipidName
        
        if(newc2TopRT.has_key(newLipidName)):
            for item in v:
                newc2TopRT[newLipidName].append(item)
        else:
            newc2TopRT[newLipidName] = v
        
        if(alias.has_key(newLipidName)):
            alias[newLipidName].add(n)
        else:
            alias[newLipidName] = set([n])
            
    
    for (n, v) in newc2TopRT.iteritems():
        if(len(v) == 0):
            newc2TopRTAvg[n] = 0.0
        else:
            newc2TopRTAvg[n] = sum(v) / len(v)
            
    
    return (newc2TopRTAvg, newc2TopRT, alias)

def genFullName(newLipidName, aliasNames):
    aliasList = sorted(list(aliasNames))
            
#     partFullName = ""
#     for n in aliasList:
#         partFullName = n + "/" + partFullName
        
    
    partFullName = aliasList[0] + "/"    
    newFullName = partFullName + newLipidName 
    
    #只要被调用一次，就需要保存一次新名字到原始名字的映射
    if(lipidAlias.has_key(newFullName)):
        lipidAlias[newFullName].add(aliasList[0])
    else:
        lipidAlias[newFullName] = set([aliasList[0]])
    
    return newFullName


def p2dot4c(f2c, c2TopRTAvg, c2TopRT, c2Grade):
    logger.debug("Start to process lipid with grade values as A,B,C,D")
    vldLipidInfo = []
    
    (newc2TopRTAvg, newc2TopRT, alias) = reorderAllGradeABCD(c2TopRT, c2Grade)
    
#     print("alias = " + str(alias))
    
    for (k, v) in f2c.iteritems():
        lipidNames = list(v)
        
        for n in lipidNames:
            grades = c2Grade[n]
            flag = isAllABCD(grades)
            
            if(flag == False):
                continue
            
            newLipidName = reduceName(n)
            
            
            #保存需要重新计算平均值的topRT值
            mergeValues = []
            #保存需要单独保存的topRT值
            unMergeValues = []
            
            topRTs = newc2TopRT[newLipidName]
            lipidAvg = newc2TopRTAvg[newLipidName]
            
            for topRT in topRTs:
                diff = math.fabs(topRT - lipidAvg)
                
                if(diff <= topRTRange):
                    mergeValues.append(topRT)
                else:
                    unMergeValues.append(topRT)
            #把合并前后的名字都组合起来
            newFullName = genFullName(newLipidName, alias[newLipidName])
            if(len(mergeValues) >= 1):
                newAvg = sum(mergeValues) / len(mergeValues)
                vldLipidInfo.append((newFullName, newAvg))
            
            for value in unMergeValues:
                vldLipidInfo.append((newFullName, value))
            
            
            
    logger.debug("End to process lipid with grade values as A,B,C,D")
    vldLipidInfo = set(vldLipidInfo) #用set去除重复的部分
    vldLipidInfo = list(vldLipidInfo)#再转回list
  
    return vldLipidInfo      

def combineLipidInfo(lipidInfoIn2dot1, lipidInfoIn2dot2, lipidInfoIn2dot3, lipidInfoIn2dot4a, lipidInfoIn2dot4b, lipidInfoIn2dot4c, c2f, c2om):
    finalLipidInfo = []
    
#     print("keys of c2f:" + str(list(c2f)))
#     print("keys of lipidAlias:" + str(list(lipidAlias)))
    
    allPairs = lipidInfoIn2dot1 + lipidInfoIn2dot2 + lipidInfoIn2dot3 + lipidInfoIn2dot4a + lipidInfoIn2dot4b + lipidInfoIn2dot4c
    
    for item in allPairs:
        lipidName = item[0]
        topRTAvg = item[1]
        
        origLipidName = item[0]
        if(lipidAlias.has_key(lipidName)):
            origLipidName = list(lipidAlias[lipidName])[0]
        
        
        
        formula = c2f[origLipidName]
        obsMz = c2om[origLipidName]
        
        lipidFinalItem = (lipidName, formula, obsMz, topRTAvg)
        finalLipidInfo.append(lipidFinalItem)
        
    
    return finalLipidInfo
        
        
        
        
    
    return finalLipidInfo

def getAreaMap2(lipidInfo):
    logger.debug("Start to get lipid name mapping to area info 2")
    
    c2AreaMap = {}
    for item in lipidInfo:
        area = item[7]
        lipidName = item[0]
        fileName = item[-1]
        
        key = (lipidName, fileName)
        if(c2AreaMap.has_key(key)):
            #c2AreaMap[key].add(area);
            c2AreaMap[key].append(area)
        else:
            c2AreaMap[key] = [area]
    logger.debug("End to get lipid name mapping to area info 2")
    
    return c2AreaMap



    

    
def getFinalC2AreaMap(lipidInfo):
    
    c2Area = getAreaMap2(lipidInfo)
    
    logger.debug("c2Area = " + str(c2Area))
    
    vldLipidAreaInfo = {}
    for item in lipidInfo:        
        lipidName = item[0]
        fileName = item[-1]  
        
        
        key = (lipidName, fileName)
            
        areas = c2Area[key]
        vldLipidAreaInfo[key] =  areas
        
#         if(vldLipidInfo.has_key(lipidName)):
#             vldLipidInfo[lipidName] = vldLipidInfo[lipidName] + areas
#         else:
#             vldLipidInfo[lipidName] =  areas      
    logger.debug("p6 result = " + str(vldLipidAreaInfo))
    
    return vldLipidAreaInfo





def p6(finalLipidInfo, vldLipidAreaInfo, dataBook):
    
    fileNames = list(dataBook)
    
    finalAreaInfo = {}
    
    for item in finalLipidInfo:
        lipidName = item[0]
        
    
        #先处理没有别名的部分
        for f in fileNames:
            key = (lipidName, f)
            if(vldLipidAreaInfo.has_key(key)):
                finalAreaInfo[key] = vldLipidAreaInfo[key]
            
        
        #处理有别名部分
        lipidNames = []
        if(lipidAlias.has_key(lipidName)):
            lipidNames = lipidAlias[lipidName]
        
        for n in lipidNames:
            for f in fileNames:
                key1 = (lipidName, f)
                key2 = (n, f)
                
                if(finalAreaInfo.has_key(key1)):                    
                    if(vldLipidAreaInfo.has_key(key2)):
#                         logger.info("Shold never be run here")
#                         logger.info("key1 = " + str(key1))
#                         logger.info("key2 = " + str(key2))
#                         logger.info("lipidNames:" + str(lipidNames))
#                         exit()
                        finalAreaInfo[key1] = finalAreaInfo[key1] + vldLipidAreaInfo[key2]
                    else:
                        finalAreaInfo[key1] = finalAreaInfo[key1]                        
                else:
                    if(vldLipidAreaInfo.has_key(key2)):
                        finalAreaInfo[key1] = vldLipidAreaInfo[key2]
       
    #取出value中重复的值       
    for(k, v) in finalAreaInfo.iteritems():
            finalAreaInfo[k] = set(v)   
        
    logger.debug("final area info = " + str(finalAreaInfo))    
    
    return finalAreaInfo


def getMSCoreMap(lipidInfo):
    logger.debug("Start to get lipid name mapping to mscore info")
    
    c2MScoreMap = {}
    for item in lipidInfo:
        mscore = item[8]
        lipidName = item[0]
        fileName = item[-1]
        
        key = (lipidName, fileName)
        if(c2MScoreMap.has_key(key)):
            #c2AreaMap[key].add(area);
            c2MScoreMap[key].append(mscore)
        else:
            c2MScoreMap[key] = [mscore]
    logger.debug("End to get lipid name mapping to mscore info")
    
    return c2MScoreMap





def getFinalC2MSCoreMap(lipidInfo):
    
    c2MSCore = getMSCoreMap(lipidInfo)
    
    logger.debug("c2MSCore = " + str(c2MSCore))
    
    vldLipidMScoreInfo = {}
    for item in lipidInfo:        
        lipidName = item[0]
        fileName = item[-1]  
        
        
        key = (lipidName, fileName)
            
        areas = c2MSCore[key]
        vldLipidMScoreInfo[key] =  areas
    
    logger.debug("p7 result = " + str(vldLipidMScoreInfo))
    
    return vldLipidMScoreInfo
                

def p7(finalLipidInfo, vldLipidMSoreInfo, dataBook):
    
    fileNames = list(dataBook)
    
    finalMScoreInfo = {}
    
    for item in finalLipidInfo:
        lipidName = item[0]
        
    
        #先处理没有别名的部分
        for f in fileNames:
            key = (lipidName, f)
            if(vldLipidMSoreInfo.has_key(key)):
                finalMScoreInfo[key] = vldLipidMSoreInfo[key]
            
        
        #处理有别名部分
        lipidNames = []
        if(lipidAlias.has_key(lipidName)):
            lipidNames = lipidAlias[lipidName]
        
        for n in lipidNames:
            for f in fileNames:
                key1 = (lipidName, f)
                key2 = (n, f)
                
                if(finalMScoreInfo.has_key(key1)):
                    if(vldLipidMSoreInfo.has_key(key2)):
                        finalMScoreInfo[key1] = finalMScoreInfo[key1] + vldLipidMSoreInfo[key2]
                    else:
                        finalMScoreInfo[key1] = finalMScoreInfo[key1]                        
                else:
                    if(vldLipidMSoreInfo.has_key(key2)):
                        finalMScoreInfo[key1] = vldLipidMSoreInfo[key2]
       
    #取出value中重复的值       
    for(k, v) in finalMScoreInfo.iteritems():
            finalMScoreInfo[k] = set(v)   
        
    logger.debug("final area info = " + str(finalMScoreInfo))    
    
    return finalMScoreInfo


def getTopRTMap(lipidInfo):
    logger.debug("Start to get lipid name mapping to toprt info")
    
    c2TopRTMap = {}
    for item in lipidInfo:
        topRT = item[2]
        lipidName = item[0]
        fileName = item[-1]
        
        key = (lipidName, fileName)
        if(c2TopRTMap.has_key(key)):
            #c2AreaMap[key].add(area);
            c2TopRTMap[key].append(topRT)
        else:
            c2TopRTMap[key] = [topRT]
    logger.debug("End to get lipid name mapping to mscore info")
    
    return c2TopRTMap





def getFinalC2TopRTMap(lipidInfo):
    
    c2TopRT = getTopRTMap(lipidInfo)
    
    logger.debug("c2TopRT = " + str(c2TopRT))
    
    vldLipidTopRTInfo = {}
    for item in lipidInfo:        
        lipidName = item[0]
        fileName = item[-1]  
        
        
        key = (lipidName, fileName)
            
        areas = c2TopRT[key]
        vldLipidTopRTInfo[key] =  areas
    
    logger.debug("p8 result = " + str(vldLipidTopRTInfo))
    
    return vldLipidTopRTInfo
                

def p8(finalLipidInfo, vldLipidTopRTInfo, dataBook):
    
    fileNames = list(dataBook)
    
    finalTopRTInfo = {}
    
    for item in finalLipidInfo:
        lipidName = item[0]
        
    
        #先处理没有别名的部分
        for f in fileNames:
            key = (lipidName, f)
            if(vldLipidTopRTInfo.has_key(key)):
                finalTopRTInfo[key] = vldLipidTopRTInfo[key]
            
        
        #处理有别名部分
        lipidNames = []
        if(lipidAlias.has_key(lipidName)):
            lipidNames = lipidAlias[lipidName]
        
        for n in lipidNames:
            for f in fileNames:
                key1 = (lipidName, f)
                key2 = (n, f)
                
                if(finalTopRTInfo.has_key(key1)):
                    if(vldLipidTopRTInfo.has_key(key2)):
                        finalTopRTInfo[key1] = finalTopRTInfo[key1] + vldLipidTopRTInfo[key2]
                    else:
                        finalTopRTInfo[key1] = finalTopRTInfo[key1]                        
                else:
                    if(vldLipidTopRTInfo.has_key(key2)):
                        finalTopRTInfo[key1] = vldLipidTopRTInfo[key2]
       
    #取出value中重复的值       
    for(k, v) in finalTopRTInfo.iteritems():
            finalTopRTInfo[k] = set(v)   
        
    logger.debug("final area info = " + str(finalTopRTInfo))    
    
    return finalTopRTInfo


            
def initWordBook():
    wb = Workbook()
    #remove default worksheet
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    
    return wb           
        
def createOutputDataBook():

    outputDataBook = initWordBook()    
    outputDataBook.create_sheet("Sheet-1", 0)
    outputDataBook.create_sheet("Sheet-2", 1)
    outputDataBook.create_sheet("Sheet-3", 2)
    
    return outputDataBook   
    
def saveOutputDataBook(wordbook, output="newbook.xlsx"):
    wordbook.save(filename = output)  


def getSortedKey(item):
    return item[0]


def getCurrTime():    
    st = time.localtime()
    year = st.tm_year
    month = st.tm_mon
    day = st.tm_mday
    hour = st.tm_hour
    miniute = st.tm_min
    sec = st.tm_sec
    
    strTime = str(year) + "-" + str("%02d" % month) + "-" + str("%02d" % day) + "-" + str("%02d" % hour) + ":" + str("%02d" % miniute) + ":" + str("%02d" % sec)
    
    return strTime

def initRangeValue(ms2WindowValue, topRTRangeValue):
    global ms2Window 
    ms2Window = ms2WindowValue
    global topRTRange
    topRTRange = topRTRangeValue;
    
def run_lipid_process(inputFolderPath, outputFilename, textField):
    
    #得到目录./lipddata下的所有文件的内容，以dict的形式组织，该dict的key是文件名，value是文件的内容，文件的内容又
    #以dict的形式进行组织，key是对应文件中的每1列的标题，value是该列对应的内容
    
    textField.insert('insert', "[%s]: Start to load files in folder: %s\n" % (getCurrTime(), inputFolderPath))
    
    textField.insert('insert', "[%s]: Initial ms2Window value is: %s\n" % (getCurrTime(), str(ms2Window)))
    
    textField.insert('insert', "[%s]: Initial topRTRange value is: %s\n" % (getCurrTime(), str(topRTRange)))
    
    textField.insert('insert', "[%s]: Start to load files in folder: %s\n" % (getCurrTime(), inputFolderPath))    
    logger.info("Start to load data");
    #dataBook = loadAllTextFile(inputFolderPath, logger)    
    dataBook = makeTestData()
    logger.info("End to load data")
    textField.insert('insert', "[%s]: End load files in folder: %s\n" % (getCurrTime(), inputFolderPath))
    
    logger.info("Start to make databook as tuple list")
    textField.insert('insert', "[%s]: Start to make databook as tuple list\n" % (getCurrTime()))
    lipidInfo = makeTuple(dataBook)
    textField.insert('insert', "[%s]: End to make databook as tuple list\n" % (getCurrTime()))
    logger.info("End to make databook as tuple list")
    
    logger.info("Start to remove item that the gap between RT and TopRT is bigger than ms2Window(0.2)")
    textField.insert('insert', "[%s]: Start to remove item that the gap between RT and TopRT is bigger than ms2Window( %f )\n" % (getCurrTime(), ms2Window))
    lipidInfo = rm0dot2(lipidInfo)
    textField.insert('insert', "[%s]: End to remove item that the gap between RT and TopRT is bigger than ms2Window( %f )\n" % (getCurrTime(), ms2Window))
    logger.info("End to remove item that the gap between RT and TopRT is bigger than ms2Window(0.2)")
    
    logger.info("Start to calculate Average of toprt values")
    textField.insert('insert',"[%s]: Start to calculate Average of toprt values\n"% (getCurrTime()))
    (c2TopRTAvg, c2TopRT) = calTopRTAvg(lipidInfo)
    textField.insert('insert',"[%s]: End to calculate Average of toprt values\n"% (getCurrTime()))
    logger.info("End to calculate Average of toprt values")
    
    logger.info("Start to calculate f->c and c->f map")
    textField.insert('insert',"[%s]: Start to calculate f->c and c->f map\n"% (getCurrTime()))
    (f2c,c2f) = getFormulaMap(lipidInfo)
    textField.insert('insert',"[%s]: End to calculate f->c and c->f map\n"% (getCurrTime()))
    logger.info("End to calculate f->c and c->f map")
    
    logger.info("Start to calculate c->grade map")
    textField.insert('insert',"[%s]: Start to calculate c->grade map\n"% (getCurrTime()))
    c2Grade = getGradeMap(lipidInfo)
    textField.insert('insert',"[%s]: End to calculate c->grade map\n"% (getCurrTime()))
    logger.info("End to calculate c->grade map")
    
    logger.info("Start to process 2dot1")
    textField.insert('insert',"[%s]: Start to process 2dot1\n"% (getCurrTime()))
    lipidInfoIn2dot1 = p2dot1(f2c, c2TopRTAvg,c2TopRT)
    textField.insert('insert',"[%s]: End to process 2dot1\n"% (getCurrTime()))
    logger.info("End to process 2dot1")
    
    logger.info("Start to process 2dot2")
    textField.insert('insert',"[%s]: Start to process 2dot2\n"% (getCurrTime()))
    lipidInfoIn2dot2 = p2dot2(f2c, c2TopRT)
    textField.insert('insert',"[%s]: End to process 2dot2\n"% (getCurrTime()))
    logger.info("End to process 2dot2")
    
    logger.info("Start to process 2dot3")
    textField.insert('insert',"[%s]: Start to process 2dot3\n"% (getCurrTime()))
    lipidInfoIn2dot3 = p2dot3(f2c, c2TopRT)
    textField.insert('insert',"[%s]: End to process 2dot3\n"% (getCurrTime()))
    logger.info("End to process 2dot3")
    

    logger.info("Start to remove lipid name in step 2.1, 2.2,2.3")
    textField.insert('insert',"[%s]: Start to remove lipid name in step 2.1, 2.2,2.3\n"% (getCurrTime()))
    rmLipidNameIn2dot123(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    textField.insert('insert',"[%s]: End to remove lipid name in step 2.1, 2.2,2.3\n"% (getCurrTime()))
    logger.info("End to remove lipid name in step 2.1, 2.2,2.3")
    
    logger.info("Start to process 2dot4a")
    textField.insert('insert',"[%s]: Start to process 2dot4a\n"% (getCurrTime()))
    lipidInfoIn2dot4a = p2dot4a(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    textField.insert('insert',"[%s]: End to process 2dot4a\n"% (getCurrTime()))
    logger.info("End to process 2dot4a")
    
    logger.info("Start to process 2dot4b")
    textField.insert('insert',"[%s]: Start to process 2dot4b\n"% (getCurrTime()))
    lipidInfoIn2dot4b = p2dot4b(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    textField.insert('insert',"[%s]: End to process 2dot4b\n"% (getCurrTime()))
    logger.info("End to process 2dot4b")
    
    logger.info("Start to process 2dot4c")
    textField.insert('insert',"[%s]: Start to process 2dot4c\n"% (getCurrTime()))
    lipidInfoIn2dot4c = p2dot4c(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    textField.insert('insert',"[%s]: End to process 2dot4c\n"% (getCurrTime()))
    logger.info("End to process 2dot4c")
    
    
    logger.info("Start to get c->obsmz map")
    textField.insert('insert',"[%s]: Start to get c->obsmz map\n"% (getCurrTime()))
    c2om = getObsMZMap(lipidInfo)
    textField.insert('insert',"[%s]: End to get c->obsmz map\n"% (getCurrTime()))
    logger.info("End to get c->obsmz map")
    #将上面所有lipidInfo[num]dot[num]的化合物归并到一起，list中的每个元素是一个元组(lipidName, formula, ObsMZ, TopRT ...)
    logger.info("Start to combine all lipind info together")
    textField.insert('insert',"[%s]: Start to combine all lipind info together\n"% (getCurrTime()))
    finalLipidInfo = combineLipidInfo(lipidInfoIn2dot1, lipidInfoIn2dot2, lipidInfoIn2dot3, lipidInfoIn2dot4a, lipidInfoIn2dot4b, lipidInfoIn2dot4c, c2f, c2om)
    textField.insert('insert',"[%s]: End to combine all lipind info together\n"% (getCurrTime()))
    logger.info("End to combine all lipind info together")
    #pprint.pprint(finalLipidInfo)
    logger.info("Start to get (c,filename) -> area map")
    textField.insert('insert',"[%s]: Start to get (c,filename) -> area map\n"% (getCurrTime()))
    c2Areas = getFinalC2AreaMap(lipidInfo)
    textField.insert('insert',"[%s]: End to get (c,filename) -> area map\n"% (getCurrTime()))
    logger.info("End to get (c,filename) -> area map")
    
    logger.info("Start to get (c,filename) -> mscore map")
    textField.insert('insert',"[%s]: Start to get (c,filename) -> mscore map\n"% (getCurrTime()))
    c2MSCores = getFinalC2MSCoreMap(lipidInfo)
    textField.insert('insert',"[%s]: End to get (c,filename) -> mscore map\n"% (getCurrTime()))
    logger.info("End to get (c,filename) -> mscore map")
    
    logger.info("Start to get (c,filename) -> toprt map")
    textField.insert('insert',"[%s]: Start to get (c,filename) -> toprt map\n"% (getCurrTime()))
    c2TopRTs = getFinalC2TopRTMap(lipidInfo)
    textField.insert('insert',"[%s]: End to get (c,filename) -> toprt map\n"% (getCurrTime()))
    logger.info("End to get (c,filename) -> toprt map")
    
    logger.info("Start to get c->mScore map")
    textField.insert('insert',"[%s]: Start to get c->mScore map\n"% (getCurrTime()))
    c2ms = getC2MScore(lipidInfo)
    textField.insert('insert',"[%s]: End to get c->mScore map\n"% (getCurrTime()))
    logger.info("c2ms = " + str(c2ms))
    logger.info("End to get c->mScore mapp")
    
    logger.info("Start to process 5")
    textField.insert('insert',"[%s]: Start to process 5\n"% (getCurrTime()))
    (p5c2msScore, p5c2MaxMSScore) = p5(finalLipidInfo,c2ms)
    textField.insert('insert',"[%s]: End to process 5\n"% (getCurrTime()))
    p5FinalResult = p5c2MaxMSScore
    logger.info("c2msScoreFinal = " + str(p5c2msScore))
    logger.info("c2MaxMSFinal = " + str(p5c2MaxMSScore))
    logger.info("End to process 5")
    
    
    
    #p6FinalResult中key是(化合物名，文件名)，value是化合物在该文件中的Area值。
    logger.info("Start to process 6")
    textField.insert('insert',"[%s]: Start to process 6\n"% (getCurrTime()))
    p6FinalResult = p6(finalLipidInfo, c2Areas, dataBook)
    textField.insert('insert',"[%s]: End to process 6\n"% (getCurrTime()))
    logger.info("End to process 6")
    
    logger.info("Start to process 7")
    textField.insert('insert',"[%s]: Start to process 7\n"% (getCurrTime()))
    p7FinalResult = p7(finalLipidInfo, c2MSCores, dataBook)
    textField.insert('insert',"[%s]: End to process 7\n"% (getCurrTime()))
    logger.info("End to process 7")
    
    logger.info("Start to process 8")
    textField.insert('insert',"[%s]: Start to process 8\n"% (getCurrTime()))
    p8FinalResult = p8(finalLipidInfo, c2TopRTs, dataBook)
    textField.insert('insert',"[%s]: End to process 8\n"% (getCurrTime()))
    logger.info("End to process 8")
    
    #现在需要做的事情是构造表头，内容为：lipid，Formula，ObsMZ，TopRT，文件名1,文件名2, .....
    logger.info("Start to sort the final lipid info")
    textField.insert('insert',"[%s]: Start to sort the final lipid info\n"% (getCurrTime()))
    finalLipidInfo.sort(cmp=None, key=getSortedKey, reverse=False)
    textField.insert('insert',"[%s]: End to sort the final lipid info\n"% (getCurrTime()))
    logger.info("End to sort the final lipid info")
    
    logger.info("Start to sort file names")
    textField.insert('insert',"[%s]: Start to sort file names\n"% (getCurrTime()))
    fileNames = sorted(list(dataBook))
    textField.insert('insert',"[%s]: End to sort file names\n"% (getCurrTime()))
    logger.info("End to sort file names")
    
    
    #logger.info("Start to write line data to output file")
    #textField.insert('insert',"[%s]: Start to write line data to output file\n"% (getCurrTime()))
    sheetHeader = ["Lipid", "Formula", "ObsMZ", "TopRT", "Max m-Score"] + fileNames
    
    
    newDataBook = createOutputDataBook()
    sheet1 = newDataBook.get_sheet_by_name("Sheet-1")
    sheet2 = newDataBook.get_sheet_by_name("Sheet-2")
    sheet3 = newDataBook.get_sheet_by_name("Sheet-3")
    sheet1.append(sheetHeader)
    sheet2.append(sheetHeader)
    sheet3.append(sheetHeader)
    
    #开始向第一个worksheet写数据
    cnt = 0
    for item in finalLipidInfo:
        lipidName = item[0]
        lineData1 = []
        lineData2 = []
        lineData3 = []
        
        lineData1.append(item[0])
        lineData2.append(item[0])
        lineData3.append(item[0])
        
        formula = "null"
        for v in item[1]:
            formula = v
            break
        lineData1.append(formula)
        lineData2.append(formula)
        lineData3.append(formula)
        
        ObsMz = "null"
        for v in item[2]:
            ObsMz = v
            break 
        
        if(ObsMz[-1] == "/"):
            ObsMz = ObsMz[0:-1]
        lineData1.append(ObsMz)
        lineData2.append(formula)
        lineData3.append(formula)
        
        topRT = item[3]
        lineData1.append(topRT)
        lineData2.append(formula)
        lineData3.append(formula)
        
        maxMScore = str(p5FinalResult[lipidName])
        lineData1.append(maxMScore)
        lineData2.append(formula)
        lineData3.append(formula)
        
        #Prepare area info in sheet1        
        for f in fileNames:
            key = (lipidName, f)
            area="null"
            if(p6FinalResult.has_key(key)):
                area = ""
                for a in p6FinalResult[key]:
            
                    area = area + a + "/"
            
            if(area[-1] == "/"):
                area = area[0:-1]    
            lineData1.append(area)
            
            
        #Prepare m-Socre info in sheet1        
        for f in fileNames:
            key = (lipidName, f)
            mscore="null"
            if(p7FinalResult.has_key(key)):
                mscore = ""
                for a in p7FinalResult[key]:
            
                    mscore = mscore + a + "/"
            
            if(mscore[-1] == "/"):
                mscore = mscore[0:-1]    
            lineData2.append(mscore)
            
            
        #Prepare toprt info in sheet3        
        for f in fileNames:
            key = (lipidName, f)
            topRT="null"
            if(p8FinalResult.has_key(key)):
                topRT = ""
                for a in p8FinalResult[key]:
            
                    topRT = topRT + a + "/"
            
            if(topRT[-1] == "/"):
                topRT = topRT[0:-1]    
            lineData3.append(topRT)
        
        sheet1.append(lineData1)
        sheet2.append(lineData2)
        sheet3.append(lineData3)
        
        if((cnt + 1) % 100 == 0 or (cnt + 1) == len(finalLipidInfo)):
            logger.info("Write line data counts:  " + str(cnt + 1) + " (" + str(len(finalLipidInfo)) + ")")
        
        cnt  = cnt + 1
        
            
    logger.info("Start to write line data to output file: " + outputFilename)
    textField.insert('insert',"[%s]: Start to write line data to output file: %s\n"% (getCurrTime(), outputFilename))
    saveOutputDataBook(wordbook=newDataBook, output=outputFilename)
    textField.insert('insert',"[%s]: End to write line data to output file: %s\n"% (getCurrTime(), outputFilename))
    logger.info("End to write line data to output file : " + outputFilename)
    
#     excelProcess = popen4("start excel D:\workspace-excelprocess-final\ExcelProcessor/lipid.xlsx")
#     print("Enter to finish")
#     import sys
#     line = sys.stdin.readline()
#     #sleep(100)
#     Popen("taskkill /F /im EXCEL.EXE",shell=True)
    
       
if __name__ == '__main__':
#     initRangeValue(0.2, 0.25)
#     run_lipid_process()
#      
#     exit()
    
    
#     newAvg = 3.0
#     compndName = "n"
#     vldLipid = [(compndName, 0.999)]
#     
#     unMergeVals = [1.2,2.2,2.4,2.6,2.8,3.0, 3.2,3.4,3.6,3.8,4.0,4.2,4.5]
#     
#     reCalUnMergeVals(vldLipid, compndName, newAvg, unMergeVals)
#     logger.info("vldLipid " + str(vldLipid))
#     
#     exit()
        
    #得到目录./lipddata下的所有文件的内容，以dict的形式组织，该dict的key是文件名，value是文件的内容，文件的内容又
    #以dict的形式进行组织，key是对应文件中的每1列的标题，value是该列对应的内容
    logger.info("Start to load data");
    dataBook = loadAllTextFile("./lipiddata", logger)    
    #dataBook = makeTestData()
    logger.info("End to load data")
    
    logger.info("Start to make databook as tuple list")
    lipidInfo = makeTuple(dataBook)
    logger.info("End to make databook as tuple list")
    
    logger.info("Start to remove item that the gap between RT and TopRT is bigger than ms2Window(0.2)")
    lipidInfo = rm0dot2(lipidInfo)
    logger.info("End to remove item that the gap between RT and TopRT is bigger than ms2Window(0.2)")
    
    logger.info("Start to calculate Average of toprt values")
    (c2TopRTAvg, c2TopRT) = calTopRTAvg(lipidInfo)
    logger.info("End to calculate Average of toprt values")
    
    logger.info("Start to calculate f->c and c->f map")
    (f2c,c2f) = getFormulaMap(lipidInfo)
    logger.info("End to calculate f->c and c->f map")
    
    logger.info("Start to calculate c->grade map")
    c2Grade = getGradeMap(lipidInfo)
    logger.info("End to calculate c->grade map")
    
    logger.info("Start to process 2dot1")
    lipidInfoIn2dot1 = p2dot1(f2c, c2TopRTAvg,c2TopRT)
    logger.info("End to process 2dot1")
    
    logger.info("Start to process 2dot2")
    lipidInfoIn2dot2 = p2dot2(f2c, c2TopRT)
    logger.info("End to process 2dot2")
    
    logger.info("Start to process 2dot3")
    lipidInfoIn2dot3 = p2dot3(f2c, c2TopRT)
    logger.info("End to process 2dot3")
    

    logger.info("Start to remove lipid name in step 2.1, 2.2,2.3")
    rmLipidNameIn2dot123(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    logger.info("End to remove lipid name in step 2.1, 2.2,2.3")
    
    logger.info("Start to process 2dot4a")
    lipidInfoIn2dot4a = p2dot4a(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    logger.info("End to process 2dot4a")
    
    logger.info("Start to process 2dot4b")
    lipidInfoIn2dot4b = p2dot4b(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    logger.info("End to process 2dot4b")
    
    logger.info("Start to process 2dot4c")
    lipidInfoIn2dot4c = p2dot4c(f2c, c2TopRTAvg, c2TopRT, c2Grade)
    logger.info("End to process 2dot4c")
    
    
    logger.info("Start to get c->obsmz map")
    c2om = getObsMZMap(lipidInfo)
    logger.info("End to get c->obsmz map")
    #将上面所有lipidInfo[num]dot[num]的化合物归并到一起，list中的每个元素是一个元组(lipidName, formula, ObsMZ, TopRT ...)
    logger.info("Start to combine all lipind info together")
    finalLipidInfo = combineLipidInfo(lipidInfoIn2dot1, lipidInfoIn2dot2, lipidInfoIn2dot3, lipidInfoIn2dot4a, lipidInfoIn2dot4b, lipidInfoIn2dot4c, c2f, c2om)
    logger.info("End to combine all lipind info together")
    #pprint.pprint(finalLipidInfo)
    logger.info("Start to get (c,filename) -> area map")
    c2Areas = getFinalC2AreaMap(lipidInfo)
    
    logger.info("End to get (c,filename) -> area map")
    
    
    
    
    
    
    logger.info("Start to get (c,filename) -> mscore map")
    c2MSCores = getFinalC2MSCoreMap(lipidInfo)
    logger.info("End to get (c,filename) -> mscore map")
    
    logger.info("Start to get (c,filename) -> toprt map")
    c2TopRTs = getFinalC2TopRTMap(lipidInfo)
    logger.info("End to get (c,filename) -> toprt map")
    
    logger.info("Start to get c->mScore mapp")
    c2ms = getC2MScore(lipidInfo)
    logger.info("c2ms = " + str(c2ms))
    logger.info("End to get c->mScore mapp")
    
    logger.info("Start to process 5")
    (p5c2msScore, p5c2MaxMSScore) = p5(finalLipidInfo,c2ms)
    p5FinalResult = p5c2MaxMSScore
    logger.info("c2msScoreFinal = " + str(p5c2msScore))
    logger.info("c2MaxMSFinal = " + str(p5c2MaxMSScore))
    logger.info("End to process 5")
    
    
    
    #p6FinalResult中key是(化合物名，文件名)，value是化合物在该文件中的Area值。
    logger.info("Start to process 6")
    p6FinalResult = p6(finalLipidInfo, c2Areas, dataBook)
    logger.info("End to process 6")
    
    logger.info("Start to process 7")
    p7FinalResult = p7(finalLipidInfo, c2MSCores, dataBook)
    logger.info("End to process 7")
    
    logger.info("Start to process 8")
    p8FinalResult = p8(finalLipidInfo, c2TopRTs, dataBook)
    logger.info("End to process 8")
    
    #现在需要做的事情是构造表头，内容为：lipid，Formula，ObsMZ，TopRT，文件名1,文件名2, .....
    logger.info("Start to sort the final lipid info")
    finalLipidInfo.sort(cmp=None, key=getSortedKey, reverse=False)
    logger.info("End to sort the final lipid info")
    
    logger.info("Start to sort file names")
    fileNames = sorted(list(dataBook))
    logger.info("End to sort file names")
    
    
    logger.info("Start to write line data to output file")
    sheetHeader = ["Lipid", "Formula", "ObsMZ", "TopRT", "Max m-Score"] + fileNames
    
    
    newDataBook = createOutputDataBook()
    sheet1 = newDataBook.get_sheet_by_name("Sheet-1")
    sheet2 = newDataBook.get_sheet_by_name("Sheet-2")
    sheet3 = newDataBook.get_sheet_by_name("Sheet-3")
    sheet1.append(sheetHeader)
    sheet2.append(sheetHeader)
    sheet3.append(sheetHeader)
    
    #开始向第一个worksheet写数据
    cnt = 0
    for item in finalLipidInfo:
        lipidName = item[0]
        lineData1 = []
        lineData2 = []
        lineData3 = []
        
        lineData1.append(item[0])
        lineData2.append(item[0])
        lineData3.append(item[0])
        
        formula = "null"
        for v in item[1]:
            formula = v
            break
        lineData1.append(formula)
        lineData2.append(formula)
        lineData3.append(formula)
        
        ObsMz = "null"
        for v in item[2]:
            ObsMz = v
            break 
        
        if(ObsMz[-1] == "/"):
            ObsMz = ObsMz[0:-1]
        lineData1.append(ObsMz)
        lineData2.append(formula)
        lineData3.append(formula)
        
        topRT = item[3]
        lineData1.append(topRT)
        lineData2.append(formula)
        lineData3.append(formula)
        
        maxMScore = str(p5FinalResult[lipidName])
        lineData1.append(maxMScore)
        lineData2.append(formula)
        lineData3.append(formula)
        
        #Prepare area info in sheet1        
        for f in fileNames:
            key = (lipidName, f)
            area="null"
            if(p6FinalResult.has_key(key)):
                area = ""
                for a in p6FinalResult[key]:
            
                    area = area + a + "/"
            
            if(area[-1] == "/"):
                area = area[0:-1]    
            lineData1.append(area)
            
            
        #Prepare m-Socre info in sheet1        
        for f in fileNames:
            key = (lipidName, f)
            mscore="null"
            if(p7FinalResult.has_key(key)):
                mscore = ""
                for a in p7FinalResult[key]:
            
                    mscore = mscore + a + "/"
            
            if(mscore[-1] == "/"):
                mscore = mscore[0:-1]    
            lineData2.append(mscore)
            
            
        #Prepare toprt info in sheet3        
        for f in fileNames:
            key = (lipidName, f)
            topRT="null"
            if(p8FinalResult.has_key(key)):
                topRT = ""
                for a in p8FinalResult[key]:
            
                    topRT = topRT + a + "/"
            
            if(topRT[-1] == "/"):
                topRT = topRT[0:-1]    
            lineData3.append(topRT)
        
        sheet1.append(lineData1)
        sheet2.append(lineData2)
        sheet3.append(lineData3)
        
        if((cnt + 1) % 100 == 0 or (cnt + 1) == len(finalLipidInfo)):
            logger.info("Write line data counts:  " + str(cnt + 1) + " (" + str(len(finalLipidInfo)) + ")")
        
        cnt  = cnt + 1
        
            
    logger.info("Start to write line data to output file")
    saveOutputDataBook(wordbook=newDataBook, output="lipid.xlsx")
    logger.info("End to write line data to output file")
    
    
    logger.info("c2Areas = " + str(c2Areas))
    
    excelProcess = popen4("start excel D:\workspace-excelprocess-final\ExcelProcessor/lipid.xlsx")
    print("Enter to finish")
    import sys
    line = sys.stdin.readline()
    #sleep(100)
    Popen("taskkill /F /im EXCEL.EXE",shell=True)
    
    
    
    
    
    
    
    
    

    
    
    pass