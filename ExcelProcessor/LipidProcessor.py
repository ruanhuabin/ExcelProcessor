# encoding: utf-8
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from Logger2 import MyLogger
import logging
import math
import re
import pprint
from subprocess import Popen
from popen2 import popen4
import glob
import time
#import operator

logger = MyLogger().getLogger()
fileHandler = logger.handlers[1]
logFilename = fileHandler.baseFilename
logFileHandle = open(logFilename, "w")

def nonBlankLines(f):
    """
    Used for romoved lines that start with #
    """
    for l in f:
        line = l.rstrip()
        if line and line[0] != '#':
            yield line

def readTextFile(fileName): 
    content = []  
    with open(fileName) as f:
        for line in nonBlankLines(f):
            content.append(line)
    
    return content

def loadTextFile(fileName):
    """
    Load text file to a dict, the key of dict is the column name, value is the column value
    """
    lines = readTextFile(fileName)
    headerLine = []
    if(len(lines) > 0):
        headerLine = lines[0].split('\t')
    
    logger.info("header number: {0}, headers: {1}".format(len(headerLine), headerLine))
    headers = []
    dataDict = {}
    for header in headerLine:
        headers.append(header)
        dataDict[header] = {}
    
    """
    Organize each line of file in lines to a list of list content, 
    """
    content = []
    for item in lines:
        item = item.split('\t')
        """
        Padding empty items to each element in content so that length of items in content is equal to number of headers
        The reason we need to do this is because we need to transpose the content in the following part 
        """
        itemLen = len(item)
        headerNum = len(headerLine)
        if(itemLen < headerNum):
            for i in xrange(0, headerNum - itemLen):
                item.append('')
        content.append(item)
    
    """
    Transpose the content:固定最外层的i，将每一行的第i个元素取出来组成一个列表，作为contentT的第i个元素，这第i个元素刚好是第i列
    """    
    contentT = [[x[i] for x in content] for i in range(len(content[0]))]
    lineNum = 0
    logger.debug("Data in column style: ")
    for item in contentT:        
        lineNum += 1
        itemLen = len(item)        
        logger.debug("%d:%d:%s" % (lineNum, itemLen, str(item)))
    
    for item in contentT:
        header = item[0]
        item.pop(0)
        dataDict[header] = item
        
    return dataDict

def createExcelFile():
    wb = Workbook()    
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.create_sheet("raw", 0)
    wb.create_sheet("result", 1)
    
    return wb

def writeDataToColumn(wordBook, sheetName, data, columnIndex, rowStartIndex):
    ws = wordBook.get_sheet_by_name(sheetName)
    
    for i in range(len(data)):
        ws.cell(row = rowStartIndex + i, column = columnIndex, value = data[i])
        
def writeDataToRow(wordBook, sheetName, data, rowIndex, columnStartIndex):
    ws = wordBook.get_sheet_by_name(sheetName)
    
    for i in range(len(data)):
        ws.cell(row = rowIndex, column = columnStartIndex + i, value = data[i])

def getColumnData(dataDict, columnName):
    columnData = dataDict[columnName]
    columnData = [columnName] + columnData
    return columnData

def getColumnDataStartWith(dataDict, columnNamePrefix):
    columnNames = dataDict.keys()
    needColumnNames = [n for n in columnNames if n.startswith(columnNamePrefix + '[')]
    needColumnNames = sorted(needColumnNames)
    allColumnData = []
    for n in needColumnNames:
        data = [v for v in dataDict[n]]
        data2 = [v if v != '' else '0' for v in data ]
        data = [n] + data2
        allColumnData.append(data)
        
    return allColumnData
    
def makeColumnDataAvg(dataDict, columnNamePrefix, columnNameAlias=''):
    columnNames = dataDict.keys()
    needColumnNames = [n for n in columnNames if n.startswith(columnNamePrefix + '[')]
    
    columnData = []
    for n in needColumnNames:
        data = [v for v in dataDict[n]]
        data2 = [v if v != '' else '0' for v in data ]
        data = [float(v) for v in data2]
        columnData.append(data)
        
    """
    Transpose the columnData so as to calculate average value more easily
    contentT = [[x[i] for x in content] for i in range(len(content[0]))]
    """   
    columnDataT = [[x[i] for x in columnData] for i in xrange(0, len(columnData[0]))]
    avgResult = [ str(sum(item) / len(item)) for item in columnDataT]
    if(columnNameAlias != ''):
        avgResult = [columnNameAlias] + avgResult
    else:
        avgResult = [columnNamePrefix] + avgResult
    
    return avgResult

def makeColumnDataMax(dataDict, columnNamePrefix, columnNameAlias=''):
    columnNames = dataDict.keys()
    needColumnNames = [n for n in columnNames if n.startswith(columnNamePrefix + '[')]
    
    columnData = []
    for n in needColumnNames:
        data = [v for v in dataDict[n]]
        data2 = [v if v != '' else '0' for v in data ]
        data = [float(v) for v in data2]
        columnData.append(data)
        
    """
    Transpose the columnData so as to calculate average value more easily
    contentT = [[x[i] for x in content] for i in range(len(content[0]))]
    """   
    columnDataT = [[x[i] for x in columnData] for i in xrange(0, len(columnData[0]))]
    maxResult = [ str(max(item)) for item in columnDataT]
    if(columnNameAlias != ''):
        maxResult = [columnNameAlias] + maxResult
    else:
        maxResult = [columnNamePrefix] + maxResult
    
    return maxResult


def makeColumnDataSubstract(dataList1, dataList2, columnName):
    substractResult = [columnName]
    for i in xrange(1, len(dataList1)):
        item = str(float(dataList1[i]) - float(dataList2[i]))
        substractResult.append(item)        

    return substractResult
        
def loadDataFromExcelFile(fileName):
    logger.info("Start loading data: " + fileName)
    
    wholeWorkBook = {}
    inputDataDict = load_workbook(fileName)
    #print "end load data %s" % time.clock()
    logger.info("Finish loading data: " + fileName)
    
    sheetNames = inputDataDict.get_sheet_names()
    #print sheetNames
    sheetNames.sort()   
    for currentSheetName in sheetNames:        
        logger.info("Start to load sheet data: " + currentSheetName)
        sheetData = inputDataDict.get_sheet_by_name(currentSheetName)
        rows = sheetData.rows;
        columns = sheetData.columns;
        rowsNum = len(rows)
        columsNum = len(columns)

        cnt = 0
        sheetRowTitleValue = []
        sheetColValue = []       
        #print "Data in ", currentSheetName, ":\n"
        for row in rows:
            for colValue in row:   
                             
                if(cnt < columsNum):
                    sheetRowTitleValue.append(colValue.value)
                else:
                    sheetColValue.append(colValue.value)
                    
                cnt = cnt + 1
        
        wholeColumData = []
        for i in range(columsNum):
            columnData = []
            for j in range(rowsNum - 1):
                index = j * columsNum + i;
                #print index
                columnData.append(sheetColValue[index])
            wholeColumData.append(columnData)
               
        
        #print wholeColumData
        
        sheetDataFinal = {}
        for i in range(columsNum):
            sheetDataFinal[sheetRowTitleValue[i]] = wholeColumData[i]
            
        #print sheetDataFinal
        
        wholeWorkBook[currentSheetName] = sheetDataFinal
        
    return wholeWorkBook    

def isRowValid(row, thresholdDict):
    return True
if __name__ == "__main__":
    logger.info("Start to extract data from lipid text file")
    dataDict = loadTextFile("lipid-data-4lines.txt")
    wb = createExcelFile()
    lipidIon = getColumnData(dataDict, "LipidIon")
    writeDataToColumn(wb, "raw", lipidIon, 1,1)
    calcMz = getColumnData(dataDict, "CalcMz")
    writeDataToColumn(wb, "raw", calcMz, 2,1)
    ionFormula = getColumnData(dataDict, "IonFormula")
    writeDataToColumn(wb, "raw", ionFormula, 3,1)
    
    
    rtData = makeColumnDataAvg(dataDict, "Rt")
    writeDataToColumn(wb, "raw", rtData, 4,1)
    topPosData = makeColumnDataAvg(dataDict, "TopPos")
    writeDataToColumn(wb, "raw", topPosData, 5,1)
    areaData = makeColumnDataAvg(dataDict, "Area", "ave Area")
    writeDataToColumn(wb, "raw", areaData, 6,1)
    areaScoreData = makeColumnDataAvg(dataDict, "AreaScore")
    writeDataToColumn(wb, "raw", areaScoreData, 7,1)
    topPosMinusRt = makeColumnDataSubstract(topPosData, rtData, "TopPos-RT")
    writeDataToColumn(wb, "raw", topPosMinusRt, 8,1)
    maxMScore = makeColumnDataMax(dataDict, "mScore", "max m-score")
    writeDataToColumn(wb, "raw", maxMScore, 9,1)    
    tScoreData = makeColumnDataAvg(dataDict, "tScore","ave tScore")
    writeDataToColumn(wb, "raw", tScoreData, 10,1)
    
    allAreasData = getColumnDataStartWith(dataDict, "Area")
    columnIndex = 11
    for area in allAreasData:
        
        writeDataToColumn(wb, "raw", area, columnIndex,1)
        columnIndex = columnIndex + 1
        
    
    allMScoresData = getColumnDataStartWith(dataDict, "mScore")
    for score in allMScoresData:
        
        writeDataToColumn(wb, "raw", score, columnIndex,1)
        columnIndex = columnIndex + 1
        
    wb.save("result.xlsx")
    
    """
    We construct data in result sheet with following initial data
    """
    """
    newDataDict = {}
    newDataDict[lipidIon[0]] = lipidIon[1:]
    newDataDict[calcMz[0]] = calcMz[1:]
    newDataDict[rtData[0]] = rtData[1:]
    newDataDict[ionFormula[0]] = ionFormula[1:]
    newDataDict[topPosData[0]] = topPosData[1:]
    newDataDict[areaData[0]] = areaData[1:]
    newDataDict[areaScoreData[0]] = areaScoreData[1:]
    newDataDict[topPosMinusRt[0]] = topPosMinusRt[1:]
    newDataDict[maxMScore[0]] = maxMScore[1:]
    newDataDict[tScoreData[0]] = tScoreData[1:]
    for area in allAreasData:
        newDataDict[area[0]] = area[1:]
    for score in allMScoresData:
        newDataDict[score[0]] = score[1:]
    
    """
    newDataList = [lipidIon, calcMz, rtData, ionFormula, topPosData, areaData, areaScoreData, topPosMinusRt, maxMScore, tScoreData]
    for area in allAreasData:
        newDataList.append(area)
    for mscore in allMScoresData:
        newDataList.append(mscore)
    
    newDataListT = [[x[i] for x in newDataList] for i in xrange(0, len(newDataList[0]))]
    newDataList = newDataListT
    logger.info("new data list: {0}".format(newDataList))
    
    
    finalDataList = [newDataList[0]]
    thresholdDict = {}
    for i in xrange(1, len(newDataList)):
        row = newDataList[i]
        isValid = isRowValid(row, thresholdDict)
        if(isValid):
            finalDataList.append(row)
            
        
        
    
    
    
    
    
    excelProcess = popen4("start excel D:\\ExcelProcessor\\ExcelProcessor/result.xlsx")
    print("Enter to finish")
    import sys
    line = sys.stdin.readline()
    #sleep(100)
    Popen("taskkill /F /im EXCEL.EXE",shell=True)